#!/usr/bin/env python

from openpyxl import Workbook

book = Workbook()
sheet = book.active

sheet['A3'] = 39
sheet['B3'] = 19

rows = [
    (88, 46),
    (89, 38),
    (23, 59),
    (56, 21),
    (24, 18),
    (34, 15)
]

for row in rows:
    sheet.append(row)
# dimensions属性返回非空单元格区域的左上角和右下角单元格
print(sheet.dimensions)
# 使用min_row和max_row属性，我们可以获得包含数据的最小和最大行
print("Minimum row: {0}".format(sheet.min_row))
print("Maximum row: {0}".format(sheet.max_row))
# 通过min_column和max_column属性，我们获得了包含数据的最小和最大列。
print("Minimum column: {0}".format(sheet.min_column))
print("Maximum column: {0}".format(sheet.max_column))

for c1, c2 in sheet[sheet.dimensions]:
    print(c1.value, c2.value)

book.save('dimensions.xlsx')

