#!/usr/bin/env python

from openpyxl import Workbook
from openpyxl.styles import Alignment

book = Workbook()
sheet = book.active
# 合并单元格
sheet.merge_cells('A1:B2')

cell = sheet.cell(row=1, column=1)
cell.value = 'Sunny day'
cell.alignment = Alignment(horizontal='center', vertical='center')

book.save('merging.xlsx')

